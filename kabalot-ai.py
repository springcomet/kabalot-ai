# %pip install openai
# %pip install pymupdf
# %pip install Pillow
# %pip install dropbox

from openai import OpenAI
import fitz  # PyMuPDF
import io
import os
from PIL import Image
import mimetypes
import base64
from io import BytesIO
import json
import dropbox
import re
from pathlib import Path
from datetime import datetime



@staticmethod
def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")


def pdf_to_base64_images(pdf_path):
    #Handles PDFs with multiple pages
    pdf_document = fitz.open(pdf_path)
    base64_images = []
    temp_image_paths = []

    total_pages = len(pdf_document)

    for page_num in range(total_pages):
        page = pdf_document.load_page(page_num)
        pix = page.get_pixmap()
        img = Image.open(io.BytesIO(pix.tobytes()))
        temp_image_path = f"temp_page_{page_num}.png"
        img.save(temp_image_path, format="PNG")
        temp_image_paths.append(temp_image_path)
        base64_image = encode_image(temp_image_path)
        base64_images.append(base64_image)

    for temp_image_path in temp_image_paths:
        os.remove(temp_image_path)

    return base64_images

def jpg_to_base64_images(jpg_file_path):
    base64_images = []
    
    # Open the JPG file
    with Image.open(jpg_file_path) as img:
        # Convert the image to a BytesIO object
        buffered = BytesIO()
        img.save(buffered, format="JPEG")
        
        # Encode the BytesIO object to base64
        img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        
        # Append the base64 string to the list
        base64_images.append(img_base64)
    
    return base64_images

def extract_invoice_data(base64_image):
    if test_config and test_config.get("mock_openai"):
        print("Using mock OpenAI response")
        j = {'Details of Services Charged': [{'City': 'הרצליה', 'Zone': 'אזור התעשיה', 'License Plate': '62-728-55', 'Start Time': '15/01/2024 08:50', 'To Time': '15/01/2024 10:14', 'Minutes': '83:19', 'Charge': '8.61'}, {'City': 'רמת גן', 'Zone': 'הבימה ת\"א חניון הבימה והיכל התרבות', 'License Plate': '91-600-11', 'Start Time': '26/12/2023 14:41', 'To Time': '26/12/2023 16:06', 'Minutes': '84:46', 'Charge': '8.90'}, {'City': 'תל-אביב', 'Zone': 'מרכז העיר 17:00-חניה חופשית או חניה בתשלום באזור', 'License Plate': '91-600-11', 'Start Time': '24/01/2024 07:57', 'To Time': '24/01/2024 09:38', 'Minutes': '38:04', 'Charge': '7.87'}], 'invoice_summary': {'total_charge': '25.38', 'date_of_invoice': '3/9/24', 'invoice_number': "1234567890", 'expense_type': 'vehicle', 'type_code':'c'}}
        return json.dumps(j, ensure_ascii=False, indent=4)
    
    
    system_prompt = f"""
    You are an OCR-like data extraction tool that extracts invoice data from PDFs.
   
    1. extract the data in this invoice, grouping data according to theme/sub groups, and then output into JSON.

    2. keep the keys and values of the JSON in the original language. use native characters for the keys and values,
    dont encode them to characters like "utf-8" or "ascii".

    3. The type of data you might encounter in the invoice includes but is not limited to: supplier information such as name and identity number, itemized charges, invoice information,
   such as invoice number, taxes such as vat amount or price before vat, and total charges etc. 

    4. If the page contains no charge data, output an empty JSON object and don't make up any data.

    5. If there are blank data fields in the invoice, include them as "null" values in the JSON object.
    
    6. If there are tables in the invoice, capture all of the rows and columns in the JSON object. 
    Even if a column is blank, include it as a key in the JSON object with a null value.
    
    7. If a row is blank denote missing fields with "null" values. 
    
    8. Don't interpolate or make up data.

    9. Please maintain the table structure of the charges, i.e. capture all of the rows and columns in the JSON object.

    11. add a last group named invoice_summary which repeats the values of the total charge, date of invoice, and invoice number. 
    these values should also be included in the relevant group and repeated in this group. 
    unlike other groups, in this group the key names should be in english as specified here.
    
    12. make sure the total_charge is a number and not a string. make sure the total charge includes vat if it is included in the invoice.
    it is unlikely that vat is not payed.
    total charge will not include vat typically but not neccessarily is tagged by "סהכ לתשלום" in the invoice
    or "סהכ לתשלום כולל מעמ".
    if the total charge does not include vat or the vat value is 0 then verify again that the total charge was extracted correctly.
    if vat was not specified inn the invoice or was 0 then specify that in the JSON object.

    13. try to deduce if the type of invoice, i.e. what is it for.
    use the following possible expense types and mark them both in english and with the code in parenthesis:
    parking (p), gas (g), other vehicle expenses (c), clothing (b), office (m), supplies and equipment (s),
    maintenance and repair (9), food and refreshment (f). If you can't deduce the expense type, mark it as "?". 
    add the expense type name and the expense type code in separate fields, expense_type and type_code respectively,
    to the invoice_summary grouped mentioned above.
    internet services such as hosting, domain registration and chatgpt should be marked as office (m).

    14. if the invoice is for a vehicle related purchase such as parking or gas only add the the amounts to the total charge 
    that are associated to vehicle with registration number 9160011 or 27228903. add the number of items actually added to the total charge
    to the invoice_summary group as the field related_items and the total number of items in the invoice regardless of registration number
    as total_items. if the ticket is not for a vehicle related purchase then add all the items to the total charge and dont add the
    related_items and the total_items fields to the invoice_summary group. 
    
    15. add the company idenity number to the invoice_summary group as the field company_id. the identy number typically starts with the digits 51
    and is tagged with the hebrew word ח.פ. in the invoice. it might also be tagged with the hebrew "עוסק מורשה".

    16. if the invoice is in a foreign currency, i.e. not in israeli shekels, add a field named currency to the invoice_summary group with the currency name.
    add the exchange rate for that currency against the israeli shekel at the date of the invoice as the field exchange_rate to the invoice_summary group..
    
    """
    
    response = client.chat.completions.create(
        model="gpt-4o",
        response_format={ "type": "json_object" },
        messages=[
            {
                "role": "system",
                "content": system_prompt
            },
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "extract the data in this invoice and output into JSON "},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{base64_image}", "detail": "high"}}
                ]
            }
        ],
        temperature=0.0,
    )
    return response.choices[0].message.content

def extract_from_multiple_pages(file_path):
    mime_type, _ = mimetypes.guess_type(file_path)
    print(f"File type: {mime_type}")
    if mime_type == 'application/pdf':
        base64_images = pdf_to_base64_images(file_path)
    elif mime_type == 'image/jpeg':
        base64_images = jpg_to_base64_images(file_path)
    else:
        raise ValueError(f"Unsupported file type: {mime_type}")

    entire_invoice = []
    for base64_image in base64_images:
        invoice_json = extract_invoice_data(base64_image)
        invoice_data = json.loads(invoice_json)
        
        # Add input file path to invoice_summary
        if 'invoice_summary' not in invoice_data:
            invoice_data['invoice_summary'] = {}
        invoice_data['invoice_summary']['input_file'] = file_path
        
        entire_invoice.append(invoice_data)
    return entire_invoice


def main_extract(config):
    for input_dir in config["input_dirs"]:
        print(f"Extracting data from files in {input_dir}")
        for filename in os.listdir(input_dir):
            print(f"Extracting data from {filename}")
            file_path = os.path.join(input_dir, filename)
            process_file(config, file_path)
    # Write Excel summary after processing all files
    write_invoice_summary_to_excel(config)

def process_file(config, file_path):
    print(f"Processing file: {file_path}")
    if os.path.isfile(file_path):
        try:
            invoice = extract_from_multiple_pages(file_path)
            link = upload_file_to_dropbox(config, file_path)
            print(invoice)
            invoice[0]['invoice_summary']['dropbox_link'] = link
            print(invoice)
            write_invoice(config, invoice)
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
    else:
        print(f"Skipping {file_path}, not a file.")

def get_safe_filename(invoice_data):
    # Handle both single invoice and list of invoices
    invoice_dict = invoice_data[0] if isinstance(invoice_data, list) else invoice_data
    
    # Safely navigate nested structure
    invoice_summary = invoice_dict.get('invoice_summary', {})
    invoice_number = invoice_summary.get('invoice_number')
    
    if not invoice_number:
        # Fallback to timestamp if no invoice number
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        invoice_number = f"invoice_{timestamp}"
    
    # Sanitize filename
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', str(invoice_number))
    return f"{safe_name}.json"

def write_invoice(config, invoice):
    output_dir = config["output_dir"]
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    filename = get_safe_filename(invoice)
    output_file = os.path.join(output_dir, filename)
    # Check if the file already exists and add a numerical postfix if necessary
    base_name, ext = os.path.splitext(output_file)
    counter = 1
    while os.path.exists(output_file):
        output_file = f"{base_name}_{counter}{ext}"
        counter += 1

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(invoice, f, ensure_ascii=False, indent=4)
    print(f"Data written to {output_file}")
    

def upload_file_to_dropbox(config, filename):
    if test_config and test_config.get("mock_dropbox"):
        print("Skipping Dropbox upload (mocked)")
        return "https://dropbox.com/link"

    # Read credentials from creds file
   
    # Initialize Dropbox client
    dbx = dropbox.Dropbox(config["dropbox_access_token"])
    print("Initialized Dropbox client.")
    
    
    # Ensure the Dropbox path exists
    # try:
    #     dbx.files_create_folder_v2(dropbox_path)
    #     print(f"Created folder {dropbox_path} in Dropbox.")
    # except dropbox.exceptions.ApiError as e:
    #     if e.error.is_path() and e.error.get_path().is_conflict():
    #         print(f"Folder {dropbox_path} already exists.")
    #     else:
    #         raise e
    # Upload files from the local directory to the Dropbox path
    if os.path.isfile(filename):
        dropbox_file_path = os.path.join(config["dropbox_path"], filename)
        print(f'Attempting to upload {filename} to Dropbox:{dropbox_file_path}...')
        with open(filename, 'rb') as f:
            dbx.files_upload(f.read(), dropbox_file_path, mode=dropbox.files.WriteMode.overwrite)
        print(f"Uploaded {filename} to {dropbox_file_path}")
        # Create a shared link for the uploaded file
        shared_link_metadata = dbx.sharing_create_shared_link_with_settings(dropbox_file_path)
        file_link = shared_link_metadata.url
        print(f"Shared link for {filename}: {file_link}")
    else:
        print(f"Skipping {filename}, not a file.")
    return file_link

def load_config(config_file):
    required_fields = {
        "output_dir": "Output directory",
        "input_dirs": "Input directories",
        "upload_path": "Upload path",
        "dropbox_path": "Dropbox path",
        "excel_path": "Excel output path"  # Changed from csv_path to excel_pathrom csv_path to excel_path
    }
    
    with open(config_file, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    # Validate all required fields exist
    missing = [field for field, name in required_fields.items() 
              if not config.get(field)]
    if missing:
        raise ValueError(f"Missing required fields: {', '.join(missing)}")
        
    return config

def get_dropbox_token(creds_file):
    """Read Dropbox access token from credentials file."""
    try:
        with open(creds_file, 'r', encoding='utf-8') as cred_file:
            creds = json.load(cred_file)
            dropbox_access_token = creds.get("dropbox")
            if not dropbox_access_token:
                raise ValueError("Dropbox access token not found in credentials file.")
            print("dropbox token: ", dropbox_access_token[-4:])
            return dropbox_access_token
    except FileNotFoundError:
        raise FileNotFoundError(f"Credentials file not found: {creds_file}")
    except json.JSONDecodeError:
        raise ValueError(f"Invalid JSON in credentials file: {creds_file}")

def write_invoice_summary_to_csv(config, invoice_data):
    """Write invoice summary data to CSV file."""
    import csv
    import os

    csv_path = config["csv_path"]
    file_exists = os.path.isfile(csv_path)
    
    try:
        # Get the invoice summary from the first page if it's a list
        invoice_dict = invoice_data[0] if isinstance(invoice_data, list) else invoice_data
        summary = invoice_dict.get('invoice_summary', {})
        
        # Define the fields we want to write
        fields = ['invoice_number', 'date_of_invoice', 'total_charge', 
                 'expense_type', 'input_file', 'dropbox_link', 'type_code']
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(csv_path), exist_ok=True)
        
        # Open file in append mode if exists, write mode if new
        mode = 'a' if file_exists else 'w'
        print(f"Writing to CSV file: {csv_path}")
        
        with open(csv_path, mode, newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            
            # Write header only if file is new
            if not file_exists:
                writer.writeheader()
                print("Wrote CSV header")
            
            # Extract only the fields we want
            row = {field: summary.get(field, '') for field in fields}
            writer.writerow(row)
            print(f"Wrote invoice summary for invoice {row.get('invoice_number', 'unknown')}")
            
    except Exception as e:
        print(f"Error writing to CSV: {str(e)}")
        raise

from openpyxl import Workbook, load_workbook

def write_invoice_summary_to_excel(config):
    """Aggregate JSON invoice data and write to an Excel file with a pivot table using Excel's PIVOTBY formula."""
    
    excel_path = config["excel_path"]
    json_dir = config["output_dir"]
    
    print(f"\nStarting Excel summary generation...")
    print(f"Reading JSON files from: {json_dir}")
    print(f"Writing to Excel file: {excel_path}")
    
    # Define the fields
    fields = ['type_code','total_charge','company_id', 'invoice_number', 'date_of_invoice', 'input_file', 'dropbox_link', 'expense_type']
    
    # Check if file exists and load it, otherwise create a new workbook
    if os.path.exists(excel_path):
        print(f"Loading existing Excel file: {excel_path}")
        wb = load_workbook(excel_path)
        ws = wb["Invoice Summary"] if "Invoice Summary" in wb.sheetnames else wb.active
        print(f"Using worksheet: {ws.title}")
    else:
        print("Creating new Excel workbook")
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Summary"
        ws.append(fields)
        print("Added header row with fields")
    
    # Find the next empty row
    next_row = ws.max_row + 1
    print(f"Starting data write from row: {next_row}")
    
    # Count files for progress tracking
    json_files = [f for f in os.listdir(json_dir) if f.endswith(".json")]
    total_files = len(json_files)
    processed_files = 0
    
    print(f"\nFound {total_files} JSON files to process")
    
    # Collect and write all JSON data
    for file in json_files:
        processed_files += 1
        print(f"\nProcessing file {processed_files}/{total_files}: {file}")
        
        file_path = os.path.join(json_dir, file)
        with open(file_path, 'r', encoding='utf-8') as f:
            invoice_data = json.load(f)
            invoice_dict = invoice_data[0] if isinstance(invoice_data, list) else invoice_data
            summary = invoice_dict.get('invoice_summary', {})
            
            # Convert total_charge to float
            total_charge = summary.get('total_charge', '')
            if isinstance(total_charge, str) and total_charge:
                try:
                    total_charge = float(total_charge.replace(',', ''))
                    summary['total_charge'] = total_charge
                    print(f"Converted total charge: {total_charge}")
                except ValueError:
                    print(f"Warning: Could not convert total charge to float: {total_charge}")
            
            row_data = [summary.get(field, '') for field in fields]
            ws.append(row_data)
            print(f"Added row for invoice: {summary.get('invoice_number', 'unknown')}")
    
    print("\nFormatting total_charge column as currency")
    # Format the total_charge column (column C) as currency
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)  # Column C (total_charge)
        cell.number_format = '#,##0.00₪'
    
    print("\nCreating Pivot Summary sheet")
    # Create Pivot Table on a new sheet
    pivot_ws = wb["Pivot Summary"] if "Pivot Summary" in wb.sheetnames else wb.create_sheet(title="Pivot Summary")
    pivot_ws.append(["Expense Type Code", "Total Amount Paid"])
    
    # Determine the exact data range
    formula = f"=PIVOTBY('Invoice Summary'!G2:G{ws.max_row}, , 'Invoice Summary'!C2:C{ws.max_row}, LAMBDA(x, SUM(x)))"
    pivot_ws["B2"] = formula
    print(f"Added pivot formula: {formula}")
    
    # Save the workbook
    wb.save(excel_path)
    print(f"\nExcel file successfully updated: {excel_path}")
    print(f"Processed {processed_files} files")
    print("Excel summary generation complete")

def test_extract(config):
    if test_config.get("test_files"):
        print("Processing test files")
        for test_file in test_config["test_files"]:
            process_file(config, test_file)
        # Write Excel summary after processing all files
        write_invoice_summary_to_excel(config)

def clean_output_directory(config):
    print("Cleaning output directory")
    output_dir = config["output_dir"]
    if os.path.exists(output_dir):
        for filename in os.listdir(output_dir):
            file_path = os.path.join(output_dir, filename)
            os.remove(file_path)
        print(f"Cleaned {output_dir}")


# Example usage
api_key = os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=api_key)
config_file = './config/kabalot.json'
creds_file='./secrets/secrets.json'

# Read JSON file to get the Dropbox path
config = load_config(config_file)
dropbox_access_token = get_dropbox_token(creds_file)
print("config: ", config)
config["dropbox_access_token"] = dropbox_access_token

test_config = {
     "test_files": [
    #     r"C:\Users\aviv\source\repos\kabalot-ai\in - Copy\IMG-20230401-WA0003.jpg",
    #     r"C:\Users\aviv\source\repos\kabalot-ai\in\38aaca37-3357-441c-94f5-6d1d051c8979_195011215Sign.pdf"],
    r"C:\Users\aviv\source\repos\kabalot-ai\in\2024-10-26 12-11.pdf",
    r"C:\Users\aviv\source\repos\kabalot-ai\in\2024-10-26 12-11.pdf"],
    "mock_openai": False,
    "mock_dropbox": True,
    "clean_output": True
}

if test_config.get("clean_output"):
    clean_output_directory(config)

if test_config.get("test_files"):
    test_extract(config)
else:
    main_extract(config)
